from flask import Flask, render_template, request, redirect, url_for
import openpyxl

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return "No file uploaded.", 400

    file = request.files['file']

    if file.filename == '':
        return "No selected file.", 400

    if file:
        inv_file = openpyxl.load_workbook(file)
        product_list = inv_file["Sheet1"]

        products_per_supplier = {}
        total_value_per_supplier = {}

        for product_row in range(2, product_list.max_row + 1):
            supplier_name = product_list.cell(product_row, 4).value
            inventory = product_list.cell(product_row, 2).value
            price = product_list.cell(product_row, 3).value

            # Number of products per supplier
            if supplier_name in products_per_supplier:
                products_per_supplier[supplier_name] += 1
            else:
                products_per_supplier[supplier_name] = 1

            # Total value per supplier
            if supplier_name in total_value_per_supplier:
                total_value_per_supplier[supplier_name] += inventory * price
            else:
                total_value_per_supplier[supplier_name] = inventory * price

        return render_template(
            'results.html',
            products_per_supplier=products_per_supplier,
            total_value_per_supplier=total_value_per_supplier,
        )

if __name__ == '__main__':
    app.run(debug=True)
